"""Importa la planilla del índice de salmos y regenera src/data/psalmIndex.data.ts.

Lee las hojas 'Año A', 'Año B', 'Año C' de la planilla (columna B = celebración,
C = página, D = antífona) y escribe el mapa que consume la app.

Uso: .venv/Scripts/python.exe scripts/import-psalm-index.py [ruta_al_xlsx]
     (por defecto: C:/Users/gusta/Downloads/Indice-Salmos-StellaMaris.xlsx)
"""
import json
import sys
from openpyxl import load_workbook

XLSX = sys.argv[1] if len(sys.argv) > 1 else r"C:/Users/gusta/Downloads/Indice-Salmos-StellaMaris.xlsx"
OUT = "src/data/psalmIndex.data.ts"

wb = load_workbook(XLSX, data_only=True)
result = {"A": {}, "B": {}, "C": {}}
counts = {}

for letter in ("A", "B", "C"):
    sheet = f"Año {letter}"
    if sheet not in wb.sheetnames:
        continue
    ws = wb[sheet]
    n = 0
    for row in ws.iter_rows(min_row=5, values_only=True):
        # columnas: 1 Sección, 2 Celebración, 3 Página, 4 Antífona, 5 Notas
        celebration = (row[1] or "").strip() if isinstance(row[1], str) else row[1]
        if not celebration or str(celebration).startswith("Otra:"):
            continue
        page = row[2]
        antiphon = row[3]
        entry = {}
        if isinstance(page, (int, float)) and page:
            entry["page"] = int(page)
        elif isinstance(page, str) and page.strip().isdigit():
            entry["page"] = int(page.strip())
        if isinstance(antiphon, str) and antiphon.strip():
            entry["antiphon"] = antiphon.strip()
        if entry:
            result[letter][str(celebration).strip()] = entry
            n += 1
    counts[letter] = n


def ts_obj(d):
    if not d:
        return "{}"
    lines = []
    for cel, entry in d.items():
        parts = []
        if "page" in entry:
            parts.append(f"page: {entry['page']}")
        if "antiphon" in entry:
            parts.append(f"antiphon: {json.dumps(entry['antiphon'], ensure_ascii=False)}")
        lines.append(f"    {json.dumps(cel, ensure_ascii=False)}: {{ {', '.join(parts)} }},")
    return "{\n" + "\n".join(lines) + "\n  }"


ts = (
    "// ARCHIVO GENERADO por scripts/import-psalm-index.py — NO editar a mano.\n"
    "// Se regenera desde la planilla del índice de salmos (Indice-Salmos-StellaMaris.xlsx).\n"
    "export interface PsalmEntryData { page?: number; antiphon?: string; }\n\n"
    "export const PSALM_INDEX_DATA: Record<'A' | 'B' | 'C', Record<string, PsalmEntryData>> = {\n"
    f"  A: {ts_obj(result['A'])},\n"
    f"  B: {ts_obj(result['B'])},\n"
    f"  C: {ts_obj(result['C'])},\n"
    "};\n"
)

with open(OUT, "w", encoding="utf-8") as f:
    f.write(ts)

print(f"Generado {OUT}")
print("Entradas por ciclo:", counts)
