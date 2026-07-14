"""Genera la plantilla de ÍNDICE DE SALMOS (una hoja por año litúrgico A/B/C).

El libro tiene un salmo por página; esta plantilla mapea cada celebración a su número
de página en el PDF de ese año. Se rellena a mano leyendo el libro (columna amarilla).

Requiere openpyxl:  .venv/Scripts/python.exe -m pip install openpyxl
Uso:               .venv/Scripts/python.exe scripts/build-indice-salmos-xlsx.py
Salida:            C:/Users/gusta/Downloads/Indice-Salmos-StellaMaris.xlsx
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = r"C:/Users/gusta/Downloads/Indice-Salmos-StellaMaris.xlsx"

INK = "1F2430"; BLUE = "2D4A7A"; GOLD = "9A7636"
input_fill = PatternFill("solid", fgColor="FFF2B2")
sec_fill = PatternFill("solid", fgColor="E7EEF7")
head_fill = PatternFill("solid", fgColor="ECEADF")
title_font = Font(size=15, bold=True, color=INK)
sub_font = Font(size=10, italic=True, color="4A4F5C")
sec_font = Font(size=11, bold=True, color=BLUE)
head_font = Font(size=10, bold=True, color=INK)
thin = Side(style="thin", color="D6D3C8")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# Estructura de celebraciones (misma para A/B/C; la PÁGINA cambia por año).
TEMPORAL = [
    ("Adviento", ["1.º Domingo de Adviento", "2.º Domingo de Adviento",
                  "3.º Domingo de Adviento", "4.º Domingo de Adviento"]),
    ("Navidad", ["Natividad del Señor (Navidad)", "Sagrada Familia",
                 "Santa María, Madre de Dios", "2.º Domingo después de Navidad",
                 "Epifanía del Señor", "Bautismo del Señor"]),
    ("Tiempo Ordinario", [f"{n}.º Domingo del Tiempo Ordinario" for n in range(2, 34)]
        + ["34.º Domingo T.O. — Cristo Rey"]),
    ("Cuaresma", ["1.º Domingo de Cuaresma", "2.º Domingo de Cuaresma",
                  "3.º Domingo de Cuaresma", "4.º Domingo de Cuaresma",
                  "5.º Domingo de Cuaresma", "Domingo de Ramos"]),
    ("Pascua", ["Domingo de Resurrección", "2.º Domingo de Pascua",
                "3.º Domingo de Pascua", "4.º Domingo de Pascua",
                "5.º Domingo de Pascua", "6.º Domingo de Pascua",
                "7.º Domingo de Pascua", "Pentecostés"]),
    ("Tras Pentecostés", ["Santísima Trinidad", "Santísimo Cuerpo y Sangre (Corpus)"]),
]
SOLEMNIDADES = [
    ("Solemnidades y fiestas (capítulo aparte)", [
        "Inmaculada Concepción", "San José", "Anunciación del Señor",
        "Ascensión del Señor", "Sagrado Corazón de Jesús",
        "Natividad de San Juan Bautista", "San Pedro y San Pablo",
        "Transfiguración del Señor", "Asunción de la Virgen María",
        "Exaltación de la Santa Cruz", "Todos los Santos",
        "Conmemoración de los Difuntos", "Presentación del Señor",
        "Otra: ____________", "Otra: ____________", "Otra: ____________",
    ]),
]


def build_sheet(ws, year_letter):
    ws.merge_cells("A1:D1")
    ws["A1"] = f"Índice de Salmos — Año {year_letter}"
    ws["A1"].font = title_font
    ws.merge_cells("A2:D2")
    ws["A2"] = ("Anota en la columna PÁGINA (amarilla) el número de página del PDF de este "
                "año donde está el salmo de cada celebración. Un salmo por página. Deja en "
                "blanco las que el libro no traiga.")
    ws["A2"].font = sub_font
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[2].height = 30

    r = 4
    for col, name in enumerate(["Sección", "Celebración", "Página", "Notas"], start=1):
        c = ws.cell(row=r, column=col, value=name)
        c.font = head_font; c.fill = head_fill; c.border = border
        c.alignment = Alignment(horizontal="center")
    r += 1

    for groups in (TEMPORAL, SOLEMNIDADES):
        for section, items in groups:
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
            sc = ws.cell(row=r, column=1, value=section)
            sc.font = sec_font; sc.fill = sec_fill; sc.border = border
            sc.alignment = Alignment(horizontal="left", vertical="center")
            r += 1
            for item in items:
                ws.cell(row=r, column=1, value="").border = border
                cel = ws.cell(row=r, column=2, value=item); cel.border = border
                pag = ws.cell(row=r, column=3, value=None)
                pag.border = border; pag.fill = input_fill
                pag.alignment = Alignment(horizontal="center")
                ws.cell(row=r, column=4, value="").border = border
                r += 1

    for col, w in zip("ABCD", [4, 42, 12, 26]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A5"


wb = Workbook()
# Hoja de instrucciones
ins = wb.active
ins.title = "Instrucciones"
ins["A1"] = "Índice de Salmos — Stella Maris"
ins["A1"].font = title_font
notes = [
    "Objetivo: mapear cada celebración a la página de su salmo en el PDF de cada año (A, B, C).",
    "El libro tiene UN salmo por página, así que basta el número de página (sin rangos).",
    "Hay una hoja por año: 'Año A', 'Año B', 'Año C'. Abre el PDF de ese año y completa su hoja.",
    "Solo llena la columna PÁGINA (amarilla). Deja en blanco lo que el libro no traiga.",
    "Las solemnidades/fiestas van en su sección (el capítulo aparte del final del libro).",
    "Este salmo musicalizado es SOLO para el coro; no se muestra al Pueblo fiel.",
    "Cuando termines, envíamelo y con eso construyo la función (después del freeze/QA).",
]
row = 3
for n in notes:
    ins.merge_cells(start_row=row, start_column=1, end_row=row, end_column=1)
    c = ins.cell(row=row, column=1, value="•  " + n)
    c.alignment = Alignment(wrap_text=True, vertical="center")
    ins.row_dimensions[row].height = 26
    row += 1
ins.column_dimensions["A"].width = 95

for letter in ("A", "B", "C"):
    build_sheet(wb.create_sheet(f"Año {letter}"), letter)

wb.save(OUT)
print("XLSX generado:", OUT)
