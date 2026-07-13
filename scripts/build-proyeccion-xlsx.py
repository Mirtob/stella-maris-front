"""Genera una hoja de cálculo EDITABLE con la proyección de ingresos del canal de YouTube.

Requiere openpyxl:  .venv/Scripts/python.exe -m pip install openpyxl
Uso:               .venv/Scripts/python.exe scripts/build-proyeccion-xlsx.py

Salida: C:/Users/gusta/Downloads/Proyeccion-YouTube-StellaMaris.xlsx
Las celdas AMARILLAS son editables (supuestos, vistas/mes, RPM); el resto son fórmulas
que recalculan solas al abrir en Excel o Google Sheets.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = r"C:/Users/gusta/Downloads/Proyeccion-YouTube-StellaMaris.xlsx"

wb = Workbook()
ws = wb.active
ws.title = "Proyección YouTube"

# ── Estilos ───────────────────────────────────────────────────────────────
INK = "1F2430"; GOLD = "9A7636"; BLUE = "2D4A7A"
input_fill = PatternFill("solid", fgColor="FFF2B2")   # amarillo = editable
head_fill = PatternFill("solid", fgColor="ECEADF")
band_fill = PatternFill("solid", fgColor="F4F2EA")
title_font = Font(name="Calibri", size=16, bold=True, color=INK)
sub_font = Font(name="Calibri", size=10, italic=True, color="4A4F5C")
sec_font = Font(name="Calibri", size=12, bold=True, color=BLUE)
head_font = Font(name="Calibri", size=10, bold=True, color=INK)
bold = Font(name="Calibri", size=10, bold=True)
thin = Side(style="thin", color="D6D3C8")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
money = '#,##0'
money2 = '#,##0.0'


def cell(ref, value=None, *, font=None, fill=None, fmt=None, align=None, bd=False):
    c = ws[ref]
    if value is not None:
        c.value = value
    if font: c.font = font
    if fill: c.fill = fill
    if fmt: c.number_format = fmt
    if align: c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=(align == "left"))
    if bd: c.border = border
    return c


# ── Encabezado ────────────────────────────────────────────────────────────
ws.merge_cells("A1:G1")
cell("A1", "Proyección de ingresos — Canal YouTube · Stella Maris", font=title_font)
ws.merge_cells("A2:G2")
cell("A2", "Celdas AMARILLAS = editables (supuestos, vistas/mes y RPM). El resto se calcula solo. "
           "Cifras ilustrativas, no garantizadas.", font=sub_font, align="left")

# ── Supuestos ─────────────────────────────────────────────────────────────
cell("A4", "SUPUESTOS (editar)", font=sec_font)
rows_sup = [
    ("A5", "B5", "USD → CLP", 950, money),
    ("A6", "B6", "Valor UF (CLP)", 39500, money),
    ("A7", "B7", "1 UF en USD", "=B6/B5", money2),
    ("A8", "B8", "Inversión a recuperar (USD)", 2500, money),
]
for aref, bref, label, val, fmt in rows_sup:
    cell(aref, label, font=bold, bd=True)
    is_formula = isinstance(val, str) and val.startswith("=")
    cell(bref, val, fmt=fmt, bd=True, align="right",
         fill=None if is_formula else input_fill)

# ── Escenarios ────────────────────────────────────────────────────────────
# Datos iniciales: (nombre, [(vistas_mes, rpm) por año 1..3])
scenarios = [
    ("ESCENARIO CONSERVADOR", [(10000, 0.7), (30000, 0.8), (60000, 0.9)]),
    ("ESCENARIO MODERADO",    [(25000, 1.0), (80000, 1.1), (180000, 1.2)]),
    ("ESCENARIO OPTIMISTA",   [(60000, 1.2), (200000, 1.3), (500000, 1.5)]),
]

headers = ["Año", "Vistas/mes (fin de año)", "RPM (USD/1000)",
           "Ingreso anual (USD)", "Acumulado (USD)", "Acum. (UF)", "¿Recupera inversión?"]

row = 10
for name, years in scenarios:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    cell(f"A{row}", name, font=sec_font)
    row += 1
    # cabecera
    for j, h in enumerate(headers):
        c = cell(f"{get_column_letter(j+1)}{row}", h, font=head_font, fill=head_fill,
                 bd=True, align="center")
    row += 1
    first_data = row
    for k, (vistas, rpm) in enumerate(years):
        r = row + k
        cell(f"A{r}", k + 1, bd=True, align="center")
        cell(f"B{r}", vistas, fill=input_fill, fmt=money, bd=True, align="right")   # editable
        cell(f"C{r}", rpm, fill=input_fill, fmt=money2, bd=True, align="right")     # editable
        # Ingreso anual = vistas/mes * 12 * RPM / 1000
        cell(f"D{r}", f"=B{r}*12*C{r}/1000", fmt=money, bd=True, align="right")
        # Acumulado = ingreso del año + acumulado del año anterior (dentro del escenario)
        if k == 0:
            cell(f"E{r}", f"=D{r}", fmt=money, bd=True, align="right")
        else:
            cell(f"E{r}", f"=E{r-1}+D{r}", fmt=money, bd=True, align="right")
        # Acumulado en UF
        cell(f"F{r}", f"=E{r}/$B$7", fmt=money2, bd=True, align="right")
        # ¿Recupera?
        cell(f"G{r}", f'=IF(E{r}>=$B$8,"Sí","Aún no")', bd=True, align="center")
    row = first_data + len(years) + 1  # línea en blanco entre escenarios

# ── Notas ─────────────────────────────────────────────────────────────────
cell(f"A{row}", "NOTAS", font=sec_font); row += 1
notes = [
    "Monetización (YouTube Partner): requiere ≥1.000 suscriptores y ≥4.000 h de reproducción "
    "en 12 meses. Hasta entonces, ingreso por anuncios = US$0.",
    "RPM = ingreso NETO del canal por cada 1.000 visualizaciones (ya descontada la comisión de "
    "YouTube). Nicho católico/musical en español (LatAm): típicamente US$0,5–1,5.",
    "Ingreso anual = Vistas/mes × 12 × RPM ÷ 1.000.",
    "Otras vías no incluidas: membresías del canal, Súper Gracias/Súper Chat, donaciones.",
    "Reinversión: el 100% de lo recaudado se destina a recursos del proyecto (equipos de "
    "grabación, honorarios de desarrollo y de profesores de música, cursos).",
    "Cifras ilustrativas, no garantizadas: dependen de la constancia de publicación y de la "
    "audiencia. Ajusta las celdas amarillas con datos reales cuando el canal entre al programa.",
]
for n in notes:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    cell(f"A{row}", "•  " + n, align="left", fill=band_fill)
    ws.row_dimensions[row].height = 30
    row += 1

# ── Anchos de columna ─────────────────────────────────────────────────────
widths = [26, 22, 16, 18, 16, 12, 20]
for j, w in enumerate(widths):
    ws.column_dimensions[get_column_letter(j+1)].width = w
ws.freeze_panes = "A4"

wb.save(OUT)
print("XLSX generado:", OUT)
